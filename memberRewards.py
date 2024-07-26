from threading import Thread
import openpyxl
import numpy as np

path= 'Member Attendance.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
sheet= wb.active

print('Gym Membership Rewards Program')

member=input('Type Member Name: ')



if member == sheet.cell.value:
    n=sheet.cell.row

    def memberRewards(n):
        count=0
        for n in range (2, sheet.max_column):
            val=sheet.cell(row=n, column=column)
            if val.value =='x':
                count+=1
        if (count < 15 and count >= 10):
            print('Congrats! You earned a 10% off coupon to use on snacks & bev inside the club.')
        elif (count < 20 and count >= 15):
            print('Congrats! You earned a free training session.')
        elif (count >=20):
            print('Congrats! You earned $10 off your monthly payment for next month.')
        else: 
          print('You have not yet earned a reward. Keep up the good work!')

    data=np.array([[cell.value for cell in row] for row in sheet.iter_rows()])
    thread_list=data[(2, sheet.max_row), (2, sheet.max_column)]
    ans=[]
        

    for n in sheet:
       thread=threading.Thread(target=memberRewards, args=n,)
       ans.append(thread)

    for thread in thread_list:
        thread.start()
    for thread in thread_list:
        thread.join()
       
        
        

        
